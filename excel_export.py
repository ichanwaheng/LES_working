

"""Export membrane nodal coordinates / deformations to Excel (.xlsx)."""

from __future__ import annotations

from pathlib import Path
from typing import Iterable, List, Optional, Sequence, Union

import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


def _autosize(ws, max_width: float = 18.0) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        width = max(len(str(c.value)) if c.value is not None else 0 for c in col)
        ws.column_dimensions[letter].width = min(max(width + 2, 10), max_width)


def write_membrane_deformations_xlsx(
    out_path: Union[str, Path],
    times: Sequence[float],
    nodes_over_time: Sequence[np.ndarray],
    reference_nodes: np.ndarray,
    fixed: Optional[np.ndarray] = None,
    iterations: Optional[Sequence[int]] = None,
    per_step_sheets: bool = True,
) -> Path:
    """Write nodal x,y,z (and ux,uy,uz) at each time step to an Excel workbook.

    Sheets
    ------
    summary :
        One row per time step (iteration, time, peak |u|, peak |uz|).
    reference :
        Undeformed / support reference coordinates for every node.
    deformations :
        Long table: every node at every time step
        (time_step, time_s, node_id, x, y, z, ux, uy, uz, fixed).
    step_XXXX (optional) :
        One sheet per time step with columns node_id, x, y, z, ux, uy, uz.
    """
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    ref = np.asarray(reference_nodes, dtype=float)
    n_nodes = ref.shape[0]
    if fixed is None:
        fixed = np.zeros(n_nodes, dtype=bool)
    else:
        fixed = np.asarray(fixed, dtype=bool)

    if len(times) != len(nodes_over_time):
        raise ValueError("times and nodes_over_time must have the same length")

    if iterations is None:
        iterations = list(range(1, len(times) + 1))

    wb = Workbook()

    # --- summary ---------------------------------------------------------
    ws_sum = wb.active
    ws_sum.title = "summary"
    ws_sum.append(
        [
            "time_step",
            "iteration",
            "time_s",
            "max_abs_u_m",
            "max_abs_uz_m",
            "n_nodes",
        ]
    )
    for c in ws_sum[1]:
        c.font = Font(bold=True)

    # --- reference -------------------------------------------------------
    ws_ref = wb.create_sheet("reference")
    ws_ref.append(["node_id", "x0_m", "y0_m", "z0_m", "fixed"])
    for c in ws_ref[1]:
        c.font = Font(bold=True)
    for i in range(n_nodes):
        ws_ref.append(
            [
                i,
                float(ref[i, 0]),
                float(ref[i, 1]),
                float(ref[i, 2]),
                bool(fixed[i]),
            ]
        )

    # --- deformations (long) ---------------------------------------------
    ws_def = wb.create_sheet("deformations")
    ws_def.append(
        [
            "time_step",
            "iteration",
            "time_s",
            "node_id",
            "x_m",
            "y_m",
            "z_m",
            "ux_m",
            "uy_m",
            "uz_m",
            "fixed",
        ]
    )
    for c in ws_def[1]:
        c.font = Font(bold=True)

    for step_i, (t, nodes, it) in enumerate(
        zip(times, nodes_over_time, iterations), start=1
    ):
        x = np.asarray(nodes, dtype=float)
        if x.shape != ref.shape:
            raise ValueError(
                f"nodes at step {step_i} have shape {x.shape}, expected {ref.shape}"
            )
        u = x - ref
        abs_u = np.linalg.norm(u, axis=1)
        ws_sum.append(
            [
                step_i,
                int(it),
                float(t),
                float(np.max(abs_u)),
                float(np.max(np.abs(u[:, 2]))),
                n_nodes,
            ]
        )

        if per_step_sheets:
            # Excel sheet title max 31 chars
            name = f"step_{step_i:04d}"
            ws_step = wb.create_sheet(name)
            ws_step.append(
                [
                    f"time_step={step_i}; iteration={int(it)}; time_s={float(t):.6g}"
                ]
            )
            ws_step["A1"].font = Font(italic=True, color="666666")
            ws_step.append(
                ["node_id", "x_m", "y_m", "z_m", "ux_m", "uy_m", "uz_m", "fixed"]
            )
            for c in ws_step[2]:
                c.font = Font(bold=True)

        for nid in range(n_nodes):
            row = [
                step_i,
                int(it),
                float(t),
                nid,
                float(x[nid, 0]),
                float(x[nid, 1]),
                float(x[nid, 2]),
                float(u[nid, 0]),
                float(u[nid, 1]),
                float(u[nid, 2]),
                bool(fixed[nid]),
            ]
            ws_def.append(row)
            if per_step_sheets:
                ws_step.append(
                    [
                        nid,
                        float(x[nid, 0]),
                        float(x[nid, 1]),
                        float(x[nid, 2]),
                        float(u[nid, 0]),
                        float(u[nid, 1]),
                        float(u[nid, 2]),
                        bool(fixed[nid]),
                    ]
                )

        if per_step_sheets:
            _autosize(ws_step)

    _autosize(ws_sum)
    _autosize(ws_ref)
    _autosize(ws_def)

    wb.save(out_path)
    return out_path


class DeformationRecorder:
    """Collect membrane node coordinates each time step for Excel export."""

    def __init__(self, reference_nodes: np.ndarray, fixed: Optional[np.ndarray] = None):
        self.reference = np.asarray(reference_nodes, dtype=float).copy()
        self.fixed = (
            np.zeros(self.reference.shape[0], dtype=bool)
            if fixed is None
            else np.asarray(fixed, dtype=bool).copy()
        )
        self.times: List[float] = []
        self.iterations: List[int] = []
        self.nodes: List[np.ndarray] = []

    def record(self, time: float, nodes: np.ndarray, iteration: Optional[int] = None) -> None:
        self.times.append(float(time))
        self.nodes.append(np.asarray(nodes, dtype=float).copy())
        self.iterations.append(
            int(iteration) if iteration is not None else len(self.times)
        )

    def write_xlsx(
        self,
        out_path: Union[str, Path],
        per_step_sheets: bool = True,
    ) -> Path:
        return write_membrane_deformations_xlsx(
            out_path,
            times=self.times,
            nodes_over_time=self.nodes,
            reference_nodes=self.reference,
            fixed=self.fixed,
            iterations=self.iterations,
            per_step_sheets=per_step_sheets,
        )


# In[ ]:





# In[ ]:




