#!/usr/bin/env python3
"""Rebuild membrane deformation GIFs from a finished QS-FSI run.

Does NOT re-run the simulation. Reads saved snapshots (and optional Excel
reference) from the output directory.

Usage
-----
    python make_deformation_gif.py
    python make_deformation_gif.py --out-dir output --fps 8 --amplify 5
    python make_deformation_gif.py --out-dir /path/to/output

In a notebook, you can also paste the ``rebuild_deformation_gifs`` call
from the ``if __name__`` block below into its own cell.
"""

from __future__ import annotations

import argparse
import os
import sys
from pathlib import Path
from typing import List, Optional, Sequence, Tuple

import numpy as np

try:
    ROOT = Path(__file__).resolve().parent
except NameError:
    ROOT = Path(os.getcwd()).resolve()

if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from viz import save_deformation_gif


def _load_from_snapshots(
    out_dir: Path,
) -> Tuple[List[np.ndarray], np.ndarray, np.ndarray, List[float]]:
    snaps = sorted(out_dir.glob("snapshot_*.npz"))
    if not snaps:
        raise FileNotFoundError(
            f"No snapshot_*.npz files in {out_dir}. "
            "Re-run the main cell once, or point --out-dir at the run folder."
        )

    nodes_over_time: List[np.ndarray] = []
    times: List[float] = []
    elements: Optional[np.ndarray] = None
    for path in snaps:
        data = np.load(path)
        nodes_over_time.append(np.asarray(data["membrane_nodes"], dtype=float))
        times.append(float(np.asarray(data["time"]).reshape(())))
        if elements is None:
            elements = np.asarray(data["membrane_elements"], dtype=int)

    assert elements is not None
    return nodes_over_time, elements, nodes_over_time[0].copy(), times


def _load_reference_from_excel(xlsx_path: Path) -> Optional[np.ndarray]:
    if not xlsx_path.is_file():
        return None
    try:
        from openpyxl import load_workbook
    except ImportError:
        return None

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    if "reference" not in wb.sheetnames:
        wb.close()
        return None
    ws = wb["reference"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()
    if not rows:
        return None
    ref = np.zeros((len(rows), 3), dtype=float)
    for i, row in enumerate(rows):
        # node_id, x0_m, y0_m, z0_m, fixed
        ref[i, 0] = float(row[1])
        ref[i, 1] = float(row[2])
        ref[i, 2] = float(row[3])
    return ref


def rebuild_deformation_gifs(
    out_dir: str | Path | None = None,
    fps: int = 8,
    amplify: float = 5.0,
    excel_name: str = "membrane_deformations.xlsx",
    nodes_over_time: Optional[Sequence[np.ndarray]] = None,
    elements: Optional[np.ndarray] = None,
    nodes0: Optional[np.ndarray] = None,
    times: Optional[Sequence[float]] = None,
) -> Tuple[Path, Optional[Path]]:
    """Build physical (+ optional amplified) deformation GIFs.

    Pass ``nodes_over_time`` / ``elements`` / ``nodes0`` / ``times`` to use
    in-memory data from the main notebook cell. Otherwise loads
    ``snapshot_*.npz`` from ``out_dir`` (and Excel reference if present).
    """
    out = Path(out_dir) if out_dir is not None else ROOT / "output"

    if nodes_over_time is not None:
        if elements is None or nodes0 is None:
            raise ValueError("elements and nodes0 are required with nodes_over_time")
        nodes_over_time = list(nodes_over_time)
        if times is None:
            times = list(range(len(nodes_over_time)))
        print(f"[GIF] using provided frames ({len(nodes_over_time)}) → {out}")
    else:
        nodes_over_time, elements, nodes0, times = _load_from_snapshots(out)
        ref = _load_reference_from_excel(out / excel_name)
        if ref is not None and ref.shape == nodes0.shape:
            nodes0 = ref
            print(f"[GIF] reference nodes from {excel_name}")
        else:
            # Fall back: keep XY from first snapshot, flatten Z to its median
            nodes0 = nodes_over_time[0].copy()
            nodes0[:, 2] = float(np.median(nodes0[:, 2]))
            print("[GIF] reference approximated from first snapshot (flat Z)")
        print(f"[GIF] loaded {len(nodes_over_time)} snapshots from {out}")

    gif_path = out / "membrane_deformation.gif"
    save_deformation_gif(
        nodes_over_time=nodes_over_time,
        elements=elements,
        nodes0=nodes0,
        out_path=gif_path,
        times=times,
        fps=fps,
        amplify=1.0,
        title="Quasi-static membrane deformation",
    )
    print(f"[GIF] deformation GIF → {gif_path}")

    gif_amp: Optional[Path] = None
    if float(amplify) != 1.0:
        gif_amp = out / "membrane_deformation_amplified.gif"
        save_deformation_gif(
            nodes_over_time=nodes_over_time,
            elements=elements,
            nodes0=nodes0,
            out_path=gif_amp,
            times=times,
            fps=fps,
            amplify=float(amplify),
            title=f"Membrane deformation (×{float(amplify):g})",
        )
        print(f"[GIF] amplified deformation GIF → {gif_amp}")

    return gif_path, gif_amp


def parse_args():
    p = argparse.ArgumentParser(description="Rebuild membrane deformation GIFs")
    p.add_argument(
        "--out-dir",
        type=str,
        default=None,
        help="Run output directory containing snapshot_*.npz (default: ./output)",
    )
    p.add_argument("--fps", type=int, default=8, help="GIF frames per second")
    p.add_argument(
        "--amplify",
        type=float,
        default=5.0,
        help="Visual scale for amplified GIF (1.0 disables it)",
    )
    args, _unknown = p.parse_known_args()
    return args


if __name__ == "__main__":
    args = parse_args()
    rebuild_deformation_gifs(
        out_dir=args.out_dir,
        fps=args.fps,
        amplify=args.amplify,
    )
